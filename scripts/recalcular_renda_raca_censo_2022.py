# -*- coding: utf-8 -*-
"""
Recalcula RENDA MEDIA e DEMOGRAFIA RACIAL por local de votacao usando os dados
ATUALIZADOS do Censo 2022 (agregados por setor censitario do IBGE), substituindo
os valores legados (race_source="geojson_legacy") nos arquivos
  resultados_geo/Censo 2022/censo_2022_XX.zip

Formula da "media da vizinhanca" do local de votacao  ==  a mesma de
scripts/gerar_setores_presidencial_2022_sc.py (build_weights):
  - para cada local, seleciona os setores mais proximos ate a CAPACIDADE adulta
    acumulada (apt16) atingir  eleitorado_local * CATCHMENT_FACTOR  (raio adaptativo,
    com piso de MIN_K_SECTORS setores e teto de MAX_RADIUS_KM);
  - peso de cada setor = apt16_setor * kernel_gaussiano(distancia), com
    sigma = max(KERNEL_FRAC * raio_local, MIN_SIGMA_KM);
  - Renda Media (local) = media ponderada de V06004 (renda media do responsavel)
    dos setores da vizinhanca;
  - Pct {Branca,Preta,Parda,Amarela,Indigena} = media ponderada da composicao
    racial (% sobre o total declarado) dos setores da vizinhanca.

Fontes IBGE (pasta "Dados IBGE"):
  - BR_setores_CD2022.gpkg                          -> geometria/centroide dos setores
  - Agregados_por_setores_demografia_BR.zip         -> apt16 (idade) e populacao
  - Agregados_por_setores_renda_responsavel_*.zip   -> V06004 (renda media responsavel)
  - Agregados_por_setores_cor_ou_raca_BR.zip        -> V01317..V01321 (cor/raca)

Coordenadas dos locais: resultados_geo/locais_votacao_2022_gpkg.zip
  (tabela locais_votacao_2022_ENRIQUECIDO), casadas ao censo por
  (nr_zona, nr_locvot, municipio normalizado) -- mesma ponte de gerar_setores.

Reexecutavel e deterministico. Requer: geopandas, pyogrio, scipy, numpy, openpyxl, pyproj.
"""

import os, io, csv, json, zipfile, sqlite3, re, unicodedata, math, time, glob, tempfile, shutil
import numpy as np
import geopandas as gpd
from scipy.spatial import cKDTree
from pyproj import Transformer

# ----------------------------------------------------------------------------- paths
PROJ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IBGE = r"C:/Users/lixov/OneDrive/Documentos/Sul Independente/Dados IBGE"
RGEO = os.path.join(PROJ, "resultados_geo")
CENSO_DIR = os.path.join(RGEO, "Censo 2022")

GPKG_SETORES = os.path.join(IBGE, "BR_setores_CD2022.gpkg")
DEMO_ZIP     = os.path.join(IBGE, "Agregados_por_setores_demografia_BR.zip")
RENDA_ZIP    = os.path.join(IBGE, "Agregados_por_setores_renda_responsavel_BR_20260508_xlsx.zip")
RACA_ZIP     = os.path.join(IBGE, "Agregados_por_setores_cor_ou_raca_BR.zip")
LOCAIS_ZIP   = os.path.join(RGEO, "locais_votacao_2022_gpkg.zip")
LOCAIS_GPKG  = os.path.join(PROJ, "scratch", "locais_gpkg", "locais_votacao_2022.gpkg")

# ----------------------------------------------------------------------------- params (== gerar_setores)
ADULT_FRAC_15_19 = 0.8
CATCHMENT_FACTOR = 2.5
KERNEL_FRAC      = 0.5
MIN_SIGMA_KM     = 0.4
MIN_K_SECTORS    = 4
MAX_RADIUS_KM    = 30.0
KQUERY           = 500     # candidatos do KDTree por local (cobre o catchment com folga)

UF_CODES = {
    "11":"RO","12":"AC","13":"AM","14":"RR","15":"PA","16":"AP","17":"TO",
    "21":"MA","22":"PI","23":"CE","24":"RN","25":"PB","26":"PE","27":"AL","28":"SE","29":"BA",
    "31":"MG","32":"ES","33":"RJ","35":"SP",
    "41":"PR","42":"SC","43":"RS","50":"MS","51":"MT","52":"GO","53":"DF",
}
UF_TO_CODE = {v: k for k, v in UF_CODES.items()}

RACE_FIELDS = [("Pct Branca","V01317"), ("Pct Preta","V01318"),
               ("Pct Amarela","V01319"), ("Pct Parda","V01320"),
               ("Pct Indigena","V01321")]

CENSO_AGE_BUCKETS = [
    "16 anos","17 anos","18 anos","19 anos","20 anos","21 a 24 anos","25 a 29 anos",
    "30 a 34 anos","35 a 39 anos","40 a 44 anos","45 a 49 anos","50 a 54 anos","55 a 59 anos",
    "60 a 64 anos","65 a 69 anos","70 a 74 anos","75 a 79 anos","80 a 84 anos","85 a 89 anos",
    "90 a 94 anos","95 a 99 anos","100 anos ou mais",
]


def log(*a):
    print(*a, flush=True)


def fnum(x, default=0.0):
    """float seguro; 'X' (sigilo IBGE), ''/None -> default."""
    if x is None:
        return default
    try:
        v = float(str(x).replace(",", "."))
        return v if math.isfinite(v) else default
    except Exception:
        return default


def norm(s):
    s = str(s or "").upper().strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


# ============================================================ IBGE: demografia (apt16, pop)
def load_demografia():
    log("[IBGE] demografia (idade/sexo) -> apt16, pop ...")
    z = zipfile.ZipFile(DEMO_ZIP)
    name = [n for n in z.namelist() if n.endswith(".csv")][0]
    out = {}
    with z.open(name) as f:
        rd = csv.reader(io.TextIOWrapper(f, encoding="latin-1"), delimiter=";")
        hdr = next(rd)
        idx = {c: i for i, c in enumerate(hdr)}
        cs_i = idx.get("CD_setor", idx.get("CD_SETOR", 0))
        cols = ("V01006","V01034","V01035","V01036","V01037","V01038","V01039","V01040","V01041")
        ci = {c: idx[c] for c in cols if c in idx}
        for row in rd:
            cs = row[cs_i]
            g = lambda c: fnum(row[ci[c]]) if c in ci else 0.0
            a1519 = ADULT_FRAC_15_19 * g("V01034")
            apt16 = (a1519 + g("V01035") + g("V01036") + g("V01037") + g("V01038")
                     + g("V01039") + g("V01040") + g("V01041"))
            out[cs] = (apt16, g("V01006"))
    log(f"        {len(out):,} setores")
    return out


# ============================================================ IBGE: renda (V06004 + V06001)
def load_renda():
    log("[IBGE] renda do responsavel (V06004) ...")
    import openpyxl
    z = zipfile.ZipFile(RENDA_ZIP)
    name = [n for n in z.namelist() if n.endswith(".xlsx")][0]
    data = io.BytesIO(z.read(name))
    wb = openpyxl.load_workbook(data, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(it)]
    idx = {c: i for i, c in enumerate(hdr)}
    cs_i = idx.get("CD_SETOR", idx.get("CD_setor", 0))
    r_i  = idx.get("V06004")
    out = {}
    for row in it:
        cs = str(row[cs_i]) if row[cs_i] is not None else ""
        if not cs:
            continue
        out[cs] = fnum(row[r_i]) if r_i is not None else 0.0
    wb.close()
    log(f"        {len(out):,} setores")
    return out


# ============================================================ IBGE: cor/raca (V01317..V01321)
def load_raca():
    log("[IBGE] cor/raca (V01317..V01321) ...")
    z = zipfile.ZipFile(RACA_ZIP)
    name = [n for n in z.namelist() if n.endswith(".csv")][0]
    cols = [v for _, v in RACE_FIELDS]
    out = {}
    with z.open(name) as f:
        rd = csv.reader(io.TextIOWrapper(f, encoding="latin-1"), delimiter=";")
        hdr = next(rd)
        idx = {c: i for i, c in enumerate(hdr)}
        cs_i = idx.get("CD_SETOR", idx.get("CD_setor", 0))
        ci = [idx[c] for c in cols]
        for row in rd:
            cs = row[cs_i]
            out[cs] = tuple(fnum(row[i]) for i in ci)   # (branca, preta, amarela, parda, indigena)
    log(f"        {len(out):,} setores")
    return out


# ============================================================ setores de uma UF (centroides + atributos)
def load_setores_uf(code, demo, renda, raca):
    gdf = gpd.read_file(GPKG_SETORES, columns=["CD_SETOR", "v0001"],
                        where=f"CD_UF='{code}' AND CD_SIT<>'9'", engine="pyogrio")
    if len(gdf) == 0:
        return None
    cent = gdf.geometry.to_crs(5880).centroid
    xy = np.column_stack([cent.x.to_numpy(), cent.y.to_numpy()])
    n = len(gdf)
    apt = np.zeros(n); rnd = np.zeros(n)
    race = np.zeros((n, 5))
    for i, cs in enumerate(gdf["CD_SETOR"].to_numpy()):
        d = demo.get(cs)
        if d:
            apt[i] = d[0] if d[0] > 0 else d[1] * 0.78
        else:
            apt[i] = fnum(gdf["v0001"].iloc[i]) * 0.78
        rnd[i] = renda.get(cs, 0.0)
        r = raca.get(cs)
        if r:
            race[i] = r
    return dict(xy=xy, apt=apt, renda=rnd, race=race, tree=cKDTree(xy))


# ============================================================ coords dos locais (gpkg ENRIQUECIDO)
def load_coords_uf(uf):
    if not os.path.exists(LOCAIS_GPKG):
        os.makedirs(os.path.dirname(LOCAIS_GPKG), exist_ok=True)
        with zipfile.ZipFile(LOCAIS_ZIP) as zl:
            g = [n for n in zl.namelist() if n.endswith(".gpkg")][0]
            with zl.open(g) as src, open(LOCAIS_GPKG, "wb") as dst:
                dst.write(src.read())
    con = sqlite3.connect(LOCAIS_GPKG); cur = con.cursor()
    cur.execute("""SELECT nr_zona,nr_locvot,nm_localidade,lat,long
                   FROM locais_votacao_2022_ENRIQUECIDO WHERE sg_uf=?""", (uf,))
    coords = {}
    for z_, l_, nm, la, lo in cur.fetchall():
        if la is None or lo is None:
            continue
        coords[(int(z_), int(l_), norm(nm))] = (float(lo), float(la))
    con.close()
    return coords


# ============================================================ formula da vizinhanca (== gerar_setores)
def neighborhood_weights(loc_xy, S, elet):
    """Retorna (idx, w) dos setores da vizinhanca do local e seus pesos
    apt16*kernel_gaussiano(dist), com raio adaptativo ao eleitorado."""
    apt = S["apt"]; nS = apt.shape[0]
    kq = min(KQUERY, nS)
    dist, idx = S["tree"].query(loc_xy, k=kq)
    dist = np.atleast_1d(dist); idx = np.atleast_1d(idx)
    order = np.argsort(dist)               # garante ordem por distancia
    dist = dist[order]; idx = idx[order]
    rmax = MAX_RADIUS_KM * 1000.0
    target = max(elet, 1.0) * CATCHMENT_FACTOR
    cumcap = np.cumsum(apt[idx])
    k = int(np.searchsorted(cumcap, target) + 1)
    k = max(k, MIN_K_SECTORS)
    k = min(k, idx.shape[0])
    sel = idx[:k]; dsel = dist[:k]
    keep = dsel <= rmax
    if keep.sum() >= MIN_K_SECTORS:
        sel = sel[keep]; dsel = dsel[keep]
    r_local = float(dsel.max()) if dsel.size else MIN_SIGMA_KM * 1000.0
    sigma = max(KERNEL_FRAC * r_local, MIN_SIGMA_KM * 1000.0)
    kern = np.exp(-(dsel * dsel) / (2 * sigma * sigma))
    w = apt[sel] * kern
    return sel, w


def compute_local(loc_xy, S, elet):
    sel, w = neighborhood_weights(loc_xy, S, elet)
    if w.sum() <= 0:
        return None
    # renda: media ponderada sobre setores com renda valida
    rnd = S["renda"][sel]
    mr = rnd > 0
    renda_val = None
    if mr.any() and w[mr].sum() > 0:
        renda_val = float(np.sum(w[mr] * rnd[mr]) / np.sum(w[mr]))
    # raca: media ponderada da composicao (%) sobre setores com populacao racial > 0
    race = S["race"][sel]                      # (k,5)
    tot = race.sum(axis=1)
    mt = tot > 0
    pct = None
    if mt.any() and w[mt].sum() > 0:
        comp = race[mt] / tot[mt][:, None]     # fracoes por setor
        wm = w[mt]
        pct = (np.sum(wm[:, None] * comp, axis=0) / wm.sum()) * 100.0  # 5 valores em %
    return renda_val, pct


# ============================================================ processa uma UF
def process_uf(code, uf, demo, renda, raca):
    zip_path = os.path.join(CENSO_DIR, f"censo_2022_{uf}.zip")
    if not os.path.exists(zip_path):
        log(f"  [{uf}] zip ausente -> pulado"); return None
    t = time.time()
    S = load_setores_uf(code, demo, renda, raca)
    if S is None:
        log(f"  [{uf}] sem setores -> pulado"); return None
    coords = load_coords_uf(uf)
    tr = Transformer.from_crs(4326, 5880, always_xy=True)

    z = zipfile.ZipFile(zip_path)
    jname = f"censo_2022_{uf}.json"
    doc = json.loads(z.read(jname).decode("utf-8"))
    resumo_name = f"censo_2022_{uf}_resumo.json"
    has_resumo = resumo_name in z.namelist()
    resumo = json.loads(z.read(resumo_name).decode("utf-8")) if has_resumo else None
    z.close()

    results = doc["RESULTS"]
    n_tot = len(results); n_match = 0; n_renda = 0; n_raca = 0
    for r in results.values():
        key = (int(r["nr_zona"]), int(r["nr_locvot"]), norm(r["nm_localidade"]))
        c = coords.get(key)
        if not c:
            continue
        n_match += 1
        lx, ly = tr.transform(c[0], c[1])
        elet = sum(fnum(r.get(f, 0)) for f in CENSO_AGE_BUCKETS)
        out = compute_local(np.array([lx, ly]), S, elet)
        if out is None:
            continue
        renda_val, pct = out
        if renda_val is not None:
            r["Renda Media"] = round(renda_val, 2); n_renda += 1
        if pct is not None:
            for (field, _), v in zip(RACE_FIELDS, pct):
                r[field] = round(float(v), 2)
            n_raca += 1

    # METADATA: marca a nova fonte/metodo
    md = doc.setdefault("METADATA", {})
    md["race_source"] = "ibge_censo2022_cor_ou_raca_setores_kernel_vizinhanca"
    md["income_source"] = "ibge_censo2022_renda_responsavel_setores_kernel_vizinhanca"
    md["renda_raca_formula"] = dict(
        metodo="media_da_vizinhanca_kernel_gaussiano (== gerar_setores)",
        catchment_factor=CATCHMENT_FACTOR, kernel_frac=KERNEL_FRAC,
        min_sigma_km=MIN_SIGMA_KM, min_k_sectors=MIN_K_SECTORS,
        max_radius_km=MAX_RADIUS_KM, adult_frac_15_19=ADULT_FRAC_15_19,
        renda_var="V06004", raca_vars=[v for _, v in RACE_FIELDS])
    if resumo is not None:
        resumo["METADATA"] = md
        resumo.setdefault("SUMMARY", {})
        resumo["SUMMARY"]["renda_raca_recalc"] = dict(
            locais=n_tot, casados=n_match, com_renda=n_renda, com_raca=n_raca)

    # reescreve o zip (atomico)
    payload = json.dumps(doc, ensure_ascii=False)
    fd, tmp = tempfile.mkstemp(suffix=".zip", dir=CENSO_DIR)
    os.close(fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zo:
        zo.writestr(jname, payload)
        if resumo is not None:
            zo.writestr(resumo_name, json.dumps(resumo, ensure_ascii=False))
    shutil.move(tmp, zip_path)

    log(f"  [{uf}] setores={S['apt'].shape[0]:,} locais={n_tot} casados={n_match} "
        f"renda={n_renda} raca={n_raca}  ({time.time()-t:.1f}s)")
    return dict(uf=uf, locais=n_tot, casados=n_match, renda=n_renda, raca=n_raca)


def main():
    import sys
    only = [a.upper() for a in sys.argv[1:]] or None
    t0 = time.time()
    demo = load_demografia()
    renda = load_renda()
    raca = load_raca()
    ufs = sorted(UF_TO_CODE.keys())
    if only:
        ufs = [u for u in ufs if u in only]
    log(f"[RUN] processando {len(ufs)} UF(s): {', '.join(ufs)}")
    summ = []
    for uf in ufs:
        s = process_uf(UF_TO_CODE[uf], uf, demo, renda, raca)
        if s:
            summ.append(s)
    log("\n=== RESUMO ===")
    tot_l = tot_m = tot_r = tot_c = 0
    for s in summ:
        tot_l += s["locais"]; tot_m += s["casados"]; tot_r += s["renda"]; tot_c += s["raca"]
    log(f"UFs={len(summ)}  locais={tot_l:,}  casados={tot_m:,}  com_renda={tot_r:,}  com_raca={tot_c:,}")
    log(f"Concluido em {time.time()-t0:.1f}s")


if __name__ == "__main__":
    main()
